import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLOR_MAP = {
    "auto_approved": "D1FAE5",
    "likely_ok":     "FEF9C3",
    "needs_review":  "FFEDD5",
    "manual_required": "FEE2E2"
}

HEADERS = ["Tag", "Type", "Manufacturer", "Model", "Quantity",
           "Capacity", "Voltage", "Confidence Score", "Review Status"]

def build_excel(result: dict, filename: str) -> bytes:
    wb = openpyxl.Workbook()

    # Sheet 1 - BOM
    ws1 = wb.active
    ws1.title = "BOM"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    for col, header in enumerate(HEADERS, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, item in enumerate(result.get("bom", []), 2):
        values = [
            item.get("equipment_tag", ""),
            item.get("equipment_type", ""),
            item.get("manufacturer", ""),
            item.get("model_number", ""),
            item.get("quantity", ""),
            item.get("capacity", ""),
            item.get("voltage", ""),
            f"{round(item.get('confidence_score', 0) * 100)}%" if item.get("confidence_score") else "",
            item.get("review_status", "")
        ]
        fill_color = COLOR_MAP.get(item.get("review_status", ""), "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.fill = fill

    for col in range(1, len(HEADERS) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 20
    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"

    # Sheet 2 - Issues
    ws2 = wb.create_sheet("Issues")
    issue_headers = ["Rule ID", "Severity", "Message", "Affected Tag"]
    for col, header in enumerate(issue_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    for row, issue in enumerate(result.get("issues", []), 2):
        ws2.cell(row=row, column=1, value=issue.get("rule_id", ""))
        ws2.cell(row=row, column=2, value=issue.get("severity", ""))
        ws2.cell(row=row, column=3, value=issue.get("message", ""))
        ws2.cell(row=row, column=4, value=issue.get("affected_tag", ""))
    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 25

    # Sheet 3 - Revision Changes (empty for now)
    ws3 = wb.create_sheet("Revision Changes")
    ws3.cell(row=1, column=1, value="Revision comparison coming in T5.2")

    # Sheet 4 - Coverage Map
    ws4 = wb.create_sheet("Coverage Map")
    ws4.cell(row=1, column=1, value="Page Number")
    ws4.cell(row=1, column=2, value="Item Count")

    # Sheet 5 - Metadata
    ws5 = wb.create_sheet("Metadata")
    metadata = [
        ("Filename", filename),
        ("Page Count", result.get("page_count", "")),
        ("Item Count", result.get("item_count", "")),
        ("Total Items Auto Approved", len([i for i in result.get("bom", []) if i.get("review_status") == "auto_approved"])),
        ("Total Items Needs Review", len([i for i in result.get("bom", []) if i.get("review_status") == "needs_review"])),
        ("Total Issues", len(result.get("issues", []))),
    ]
    for row, (label, value) in enumerate(metadata, 1):
        ws5.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws5.cell(row=row, column=2, value=value)
    for col in range(1, 3):
        ws5.column_dimensions[get_column_letter(col)].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
